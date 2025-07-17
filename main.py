from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, BackgroundTasks, Response, Request
from sqlalchemy.orm import Session
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func
from typing import List, Union, Dict
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta
from database import get_db, engine
import models
from pydantic import BaseModel, field_validator, EmailStr
from fastapi.responses import FileResponse, HTMLResponse
from openpyxl import Workbook
import tempfile
import os
import shutil
from enum import Enum
import joblib
from service.statementExtractor import extract_transactions
from dotenv import load_dotenv
import os
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit
from service.mail_service import _send_monthly_report_logic, send_email, scheduled_report_job, send_password_reset_email
from service import auth_service
from fastapi.security import OAuth2PasswordRequestForm
from fastapi import status
import jinja2

load_dotenv()

# Setup Jinja2 for email templates
template_loader = jinja2.FileSystemLoader(searchpath="./email_templates")
template_env = jinja2.Environment(loader=template_loader)

# Load the ML model
model = joblib.load('categoryFinder.pkl')

# Create database tables
models.Base.metadata.create_all(bind=engine)

app = FastAPI(openapi_tags=[
    {"name": "Authentication", "description": "Operations related to user registration and login"},
    {"name": "Expenses", "description": "Manage user expenses"},
    {"name": "Income", "description": "Manage user income records"},
    {"name": "Savings", "description": "Manage user saving records"},
    {"name": "Accounts", "description": "Manage user account balance"},
    {"name": "Summary", "description": "Monthly summary and combined analytics"},
    {"name": "Recurring Expenses", "description": "Manage user recurring expenses/subscriptions"},
    {"name": "Budget", "description": "Manage user monthly budget and savings goal"},
    {"name": "Saving Goals", "description": "Manage user saving goals"},
    {"name": "Utilities", "description": "Other utility operations like category prediction and statement upload"}
])

app.openapi_extra = {
    "security": [
        {"bearerAuth": []}
    ],
    "components": {
        "securitySchemes": {
            "bearerAuth": {
                "type": "http",
                "scheme": "bearer",
                "bearerFormat": "JWT",
                "description": "Enter your JWT Bearer token in the format 'Bearer <token>'"
            }
        }
    }
}

scheduler = BackgroundScheduler(daemon=True)
scheduler.add_job(scheduled_report_job, CronTrigger(day=1, hour=6, minute=0))
scheduler.start()
atexit.register(lambda: scheduler.shutdown())

origins = [
    "http://localhost:5173",
    "https://expense-tracker-frontend-nfhv.onrender.com",
    os.getenv("FRONTEND_URL")
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,  # Allow cookies and authorization headers
    allow_methods=["*"],  # Allow all HTTP methods (GET, POST, PUT, etc.)
    allow_headers=["*"]  # Allow all headers
)

# Pydantic models for Users and Auth
class UserBase(BaseModel):
    email: EmailStr
    name: str | None = None

class UserCreate(UserBase):
    password: str

class User(UserBase):
    id: int
    created_at: datetime | None = None
    last_login: datetime | None = None

    class Config:
        orm_mode = True

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenWithRefresh(BaseModel):
    access_token: str
    token_type: str
    # Note: refresh_token is sent as HttpOnly cookie, not in response body

class RequestPasswordResetRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

# Category mapping
CATEGORIES = {
    1: "Food",
    2: "Housing",
    3: "Transportation",
    4: "Personal",
    5: "Utility",
    6: "Recreation",
    7: "Health",
    8: "Debt"
}

INCOME_CATEGORIES = {
    1: "Salary",
    2: "Interest",
    3: "Gift",
    4: "Matured Amount",
    5: "Dividend",
    6: "Stocks",
    7: "Side Hustle",
    8: "Others"
}

SAVING_CATEGORIES = {
    1: "Stocks",
    2: "PPF",
    3: "Recurring deposit",
    4: "Fixed Deposit",
    5: "Mutual Fund",
    6: "Others",
    7: "Saving Goal",
    8: "Saving Goal Redeemed"
}

# Pydantic models for request/response
class IntentionType(str, Enum):
    NEED = "Need"
    WANT = "Want"
    SAVING = "Saving"

class ExpenseBase(BaseModel):
    date: date
    category_id: int
    amount: float
    intention: IntentionType = IntentionType.NEED
    name: str | None = None  # Make name optional with default None
    
    @field_validator("date", mode="before")
    def parse_datetime_to_date(cls, v):
        if isinstance(v, str):
            try:
                return datetime.fromisoformat(v.replace("Z", "+00:00")).date()
            except ValueError:
                pass
        elif isinstance(v, datetime):
            return v.date()
        return v

    class Config:
        orm_mode = True

class ExpenseCreate(ExpenseBase):
    pass

class Expense(ExpenseBase):
    id: int
    category: str = ""

    def __init__(self, **data):
        super().__init__(**data)
        self.category = CATEGORIES.get(self.category_id, "Unknown")

# Pydantic models for recurring expenses
class PeriodUnit(str, Enum):
    DAYS = "days"
    MONTHS = "months"
    YEARS = "years"

class PeriodBase(BaseModel):
    value: int
    unit: PeriodUnit

    @field_validator('value')
    def validate_value(cls, v):
        if not 1 <= v <= 30:
            raise ValueError('Value must be between 1 and 30')
        return v

class RecurringExpenseBase(BaseModel):
    name: str
    amount: float
    category_id: int
    intention: IntentionType = IntentionType.NEED
    subscription_period: PeriodBase = PeriodBase(value=1, unit=PeriodUnit.MONTHS)
    effective_date: date
    billing_period: PeriodBase = PeriodBase(value=1, unit=PeriodUnit.MONTHS)
    due_period: PeriodBase

    class Config:
        orm_mode = True

class RecurringExpenseCreate(RecurringExpenseBase):
    pass

class RecurringExpense(RecurringExpenseBase):
    id: int
    category: str = ""

    def __init__(self, **data):
        super().__init__(**data)
        self.category = CATEGORIES.get(self.category_id, "Unknown")

# Add this with other Pydantic models
class DueTomorrowResponse(BaseModel):
    message: str

# Add this with other Pydantic models
class DueTomorrowListResponse(BaseModel):
    subscriptions: List[RecurringExpense]

# Add this with other Pydantic models
class BudgetBase(BaseModel):
    monthly_income: float
    saving_goal: float
    category_budgets: Dict[str, float]

    @field_validator('monthly_income', 'saving_goal')
    def validate_amounts(cls, v):
        if v < 0:
            raise ValueError('Amount cannot be negative')
        return v

    @field_validator('category_budgets')
    def validate_category_budgets(cls, v):
        for category, amount in v.items():
            if amount < 0:
                raise ValueError(f'Budget for {category} cannot be negative')
        return v

class BudgetCreate(BudgetBase):
    pass

class Budget(BudgetBase):
    id: int
    created_at: datetime

    class Config:
        orm_mode = True

class CategoryPredictionRequest(BaseModel):
    name: str

# Pydantic models for Income
class IncomeBase(BaseModel):
    name: str | None = None
    amount: float
    date: date
    category_id: int
    
    @field_validator("date", mode="before")
    def parse_datetime_to_date(cls, v):
        if isinstance(v, str):
            try:
                return datetime.fromisoformat(v.replace("Z", "+00:00")).date()
            except ValueError:
                pass
        elif isinstance(v, datetime):
            return v.date()
        return v

    class Config:
        orm_mode = True

class IncomeCreate(IncomeBase):
    pass

class Income(IncomeBase):
    id: int
    category: str = ""
    created_at: datetime

    def __init__(self, **data):
        super().__init__(**data)
        self.category = INCOME_CATEGORIES.get(self.category_id, "Unknown")

# Pydantic models for Saving
class SavingBase(BaseModel):
    name: str | None = None
    amount: float
    date: date
    category_id: int
    
    @field_validator("date", mode="before")
    def parse_datetime_to_date(cls, v):
        if isinstance(v, str):
            try:
                return datetime.fromisoformat(v.replace("Z", "+00:00")).date()
            except ValueError:
                pass
        elif isinstance(v, datetime):
            return v.date()
        return v

    class Config:
        orm_mode = True

class SavingCreate(SavingBase):
    pass

class Saving(SavingBase):
    id: int
    category: str = ""
    created_at: datetime

    def __init__(self, **data):
        super().__init__(**data)
        self.category = SAVING_CATEGORIES.get(self.category_id, "Unknown")

# Pydantic models for Account
class AccountBase(BaseModel):
    balance: float

    @field_validator('balance')
    def validate_balance(cls, v):
        if v < 0:
            raise ValueError('Balance cannot be negative')
        return v

    class Config:
        orm_mode = True

class AccountCreate(AccountBase):
    pass

class Account(AccountBase):
    id: int
    created_at: datetime
    modified_at: datetime

@app.post("/predict-category")
def predict_category(request: CategoryPredictionRequest):
    try:
        predicted_category = model.predict([request.name])[0]
        return {"category": predicted_category}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/categories", tags=["Utilities"])
def get_all_categories():
    """Get all available categories for expenses, income, and savings"""
    return {
        "expense_categories": CATEGORIES,
        "income_categories": INCOME_CATEGORIES,
        "saving_categories": SAVING_CATEGORIES
    }

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Only .xls or .xlsx files are allowed")
    
    temp_file_path = f"temp_uploads/{file.filename}"
    os.makedirs("temp_uploads", exist_ok=True)
    
    with open(temp_file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        data = extract_transactions(temp_file_path)
        total_withdrawal = sum(item["Withdrawal"] for item in data)
        total_deposit = sum(item["Deposit"] for item in data)
        count = len(data)
        net_monthly_expenditure = total_withdrawal-total_deposit
    finally:
        os.remove(temp_file_path)  # Clean up temp file

    return {"data": data,
            "total_amount_withdrawn": total_withdrawal,
            "total_amount_deposited": total_deposit,
            "net_monthly_expenditure": net_monthly_expenditure,
            "total_transcations": count}

@app.post("/expenses/", response_model=Expense)
def create_expense(expense: ExpenseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    if expense.category_id not in CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid category ID. Must be between 1 and 9")
    
    db_expense = models.Expense(**expense.dict(), user_id=current_user.id)
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    return Expense(**db_expense.__dict__)

@app.post("/recurring-expenses/", response_model=RecurringExpense)
def create_recurring_expense(expense: RecurringExpenseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    if expense.category_id not in CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid category ID. Must be between 1 and 9")
    
    # Convert the period objects to database format
    db_expense = models.RecurringExpense(
        name=expense.name,
        amount=expense.amount,
        category_id=expense.category_id,
        intention=expense.intention,
        subscription_period_value=expense.subscription_period.value,
        subscription_period_unit=expense.subscription_period.unit,
        effective_date=expense.effective_date,
        billing_period_value=expense.billing_period.value,
        billing_period_unit=expense.billing_period.unit,
        due_period_value=expense.due_period.value,
        due_period_unit=expense.due_period.unit,
        user_id=current_user.id # Assign user_id
    )
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    
    # Convert back to response format
    return RecurringExpense(
        id=db_expense.id,
        name=db_expense.name,
        amount=db_expense.amount,
        category_id=db_expense.category_id,
        intention=db_expense.intention,
        subscription_period=PeriodBase(
            value=db_expense.subscription_period_value,
            unit=db_expense.subscription_period_unit
        ),
        effective_date=db_expense.effective_date,
        billing_period=PeriodBase(
            value=db_expense.billing_period_value,
            unit=db_expense.billing_period_unit
        ),
        due_period=PeriodBase(
            value=db_expense.due_period_value or 1,
            unit=db_expense.due_period_unit or "months"
        )
    )

@app.get("/expenses/", response_model=List[Expense])
def read_expenses(month: int = None, year: int = None, skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # If month and year are provided, filter expenses for that month
    query = db.query(models.Expense).filter(models.Expense.user_id == current_user.id)
    if month is not None and year is not None:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        query = query.filter(
            models.Expense.date >= start_date,
            models.Expense.date < end_date
        )
    # Sort by date first, then by created_at for consistent ordering
    expenses = query.order_by(models.Expense.date.desc(), models.Expense.created_at.desc()).offset(skip).limit(limit).all()
    return [Expense(**expense.__dict__) for expense in expenses]

@app.get("/expenses/total")
def get_total_expenses(month: int = None, year: int = None, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # If month and year are provided, filter expenses for that month
    query = db.query(models.Expense).filter(models.Expense.user_id == current_user.id)
    if month is not None and year is not None:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        # Get overall total for the month
        total = query.filter(
            models.Expense.date >= start_date,
            models.Expense.date < end_date
        ).with_entities(func.sum(models.Expense.amount)).scalar()
        
        # Get category-wise totals for the month
        category_totals = query.with_entities(
            models.Expense.category_id,
            func.sum(models.Expense.amount).label('total')
        ).filter(
            models.Expense.date >= start_date,
            models.Expense.date < end_date
        ).group_by(models.Expense.category_id).all()
    else:
        # Get overall total
        total = query.with_entities(func.sum(models.Expense.amount)).scalar()
        
        # Get category-wise totals
        category_totals = query.with_entities(
            models.Expense.category_id,
            func.sum(models.Expense.amount).label('total')
        ).group_by(models.Expense.category_id).all()
    
    # Convert category IDs to names and format the response
    category_breakdown = {
        CATEGORIES[cat_id]: amount for cat_id, amount in category_totals
    }
    
    return {
        "overall_total": total if total else 0,
        "category_breakdown": category_breakdown
    }

@app.delete("/expenses/{expense_id}")
def delete_expense(expense_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    expense = db.query(models.Expense).filter(models.Expense.id == expense_id, models.Expense.user_id == current_user.id).first()
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found or not authorized")
    
    message = f"Expense on {expense.date} for {CATEGORIES[expense.category_id]} amounting to {expense.amount} deleted successfully"
    db.delete(expense)
    db.commit()
    return message

@app.get("/expenses/export")
def export_expenses(db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # Get all expenses for the current user
    expenses = db.query(models.Expense).filter(models.Expense.user_id == current_user.id).all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses_"+str(date.today())
    
    # Add headers
    headers = ["ID", "Date", "Category", "Amount", "Intention", "Name"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.id)
        ws.cell(row=row, column=2, value=expense.date)
        ws.cell(row=row, column=3, value=CATEGORIES[expense.category_id])
        ws.cell(row=row, column=4, value=expense.amount)
        ws.cell(row=row, column=5, value=expense.intention)
        ws.cell(row=row, column=6, value=expense.name)
    
    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    # Return the file
    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="expenses.xlsx",
        background=None
    )

@app.get("/recurring-expenses/", response_model=List[RecurringExpense])
def get_recurring_expenses(db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    today = date.today()
    
    # Get all subscriptions for the current user where effective_date + subscription_period >= today
    subscriptions = db.query(models.RecurringExpense).filter(models.RecurringExpense.user_id == current_user.id).all()
    active_subscriptions = []
    
    for sub in subscriptions:
        # Calculate end date based on subscription period
        if sub.subscription_period_unit == "days":
            end_date = sub.effective_date + timedelta(days=sub.subscription_period_value)
        elif sub.subscription_period_unit == "months":
            end_date = sub.effective_date + relativedelta(months=sub.subscription_period_value)
        else:  # years
            end_date = sub.effective_date + relativedelta(years=sub.subscription_period_value)
            
        if end_date >= today:
            # Convert database model to Pydantic model with proper period objects
            subscription_dict = {
                "id": sub.id,
                "name": sub.name,
                "amount": sub.amount,
                "category_id": sub.category_id,
                "intention": sub.intention,
                "subscription_period": {
                    "value": sub.subscription_period_value,
                    "unit": sub.subscription_period_unit
                },
                "effective_date": sub.effective_date,
                "billing_period": {
                    "value": sub.billing_period_value,
                    "unit": sub.billing_period_unit
                },
                "due_period": {
                    "value": sub.due_period_value or 1,
                    "unit": sub.due_period_unit or "months"
                }
            }
            active_subscriptions.append(subscription_dict)
    
    return active_subscriptions

@app.post("/recurring-expenses/{subscription_id}/update-effective-date", response_model=RecurringExpense)
def update_subscription_effective_date(subscription_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    subscription = db.query(models.RecurringExpense).filter(models.RecurringExpense.id == subscription_id, models.RecurringExpense.user_id == current_user.id).first()
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found or not authorized")
    
    # Calculate new effective date based on billing period
    if subscription.billing_period_unit == "days":
        new_effective_date = subscription.effective_date + timedelta(days=subscription.billing_period_value)
    elif subscription.billing_period_unit == "months":
        new_effective_date = subscription.effective_date + relativedelta(months=subscription.billing_period_value)
    else:  # years
        new_effective_date = subscription.effective_date + relativedelta(years=subscription.billing_period_value)
    
    # Update the effective date
    subscription.effective_date = new_effective_date
    db.commit()
    db.refresh(subscription)
    
    # Convert to response format
    return {
        "id": subscription.id,
        "name": subscription.name,
        "amount": subscription.amount,
        "category_id": subscription.category_id,
        "intention": subscription.intention,
        "subscription_period": {
            "value": subscription.subscription_period_value,
            "unit": subscription.subscription_period_unit
        },
        "effective_date": subscription.effective_date,
        "billing_period": {
            "value": subscription.billing_period_value,
            "unit": subscription.billing_period_unit
        },
        "due_period": {
            "value": subscription.due_period_value or 1,
            "unit": subscription.due_period_unit or "months"
        }
    }

@app.put("/recurring-expenses/{subscription_id}", response_model=RecurringExpense)
def update_recurring_expense(subscription_id: int, expense: RecurringExpenseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_expense = db.query(models.RecurringExpense).filter(models.RecurringExpense.id == subscription_id, models.RecurringExpense.user_id == current_user.id).first()
    if not db_expense:
        raise HTTPException(status_code=404, detail="Subscription not found or not authorized")
    
    # Update the expense fields
    db_expense.name = expense.name
    db_expense.amount = expense.amount
    db_expense.category_id = expense.category_id
    db_expense.intention = expense.intention
    db_expense.subscription_period_value = expense.subscription_period.value
    db_expense.subscription_period_unit = expense.subscription_period.unit
    db_expense.effective_date = expense.effective_date
    db_expense.billing_period_value = expense.billing_period.value
    db_expense.billing_period_unit = expense.billing_period.unit
    db_expense.due_period_value = expense.due_period.value
    db_expense.due_period_unit = expense.due_period.unit
    
    db.commit()
    db.refresh(db_expense)
    
    # Convert to response format
    return {
        "id": db_expense.id,
        "name": db_expense.name,
        "amount": db_expense.amount,
        "category_id": db_expense.category_id,
        "intention": db_expense.intention,
        "subscription_period": {
            "value": db_expense.subscription_period_value,
            "unit": db_expense.subscription_period_unit
        },
        "effective_date": db_expense.effective_date,
        "billing_period": {
            "value": db_expense.billing_period_value,
            "unit": db_expense.billing_period_unit
        },
        "due_period": {
            "value": db_expense.due_period_value or 1,
            "unit": db_expense.due_period_unit or "months"
        }
    }

@app.get("/recurring-expenses/due-tomorrow", response_model=Union[DueTomorrowListResponse, DueTomorrowResponse])
def get_subscriptions_due_tomorrow(db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    today = datetime.now().date()
    
    # Get all active subscriptions for the current user
    subscriptions = db.query(models.RecurringExpense).filter(models.RecurringExpense.user_id == current_user.id).all()
    due_subscriptions = []
    
    for sub in subscriptions:
        # Calculate the alert date: effective_date + due_period - 1 day
        if sub.due_period_unit == "days":
            alert_date = sub.effective_date + timedelta(days=sub.due_period_value - 1)
        elif sub.due_period_unit == "months":
            alert_date = sub.effective_date + relativedelta(months=sub.due_period_value) - timedelta(days=1)
        else:  # years
            alert_date = sub.effective_date + relativedelta(years=sub.due_period_value) - timedelta(days=1)
        
        # Check if the alert date is today
        if alert_date == today:
            # Convert database model to Pydantic model with proper period objects
            subscription_dict = {
                "id": sub.id,
                "name": sub.name,
                "amount": sub.amount,
                "category_id": sub.category_id,
                "intention": sub.intention,
                "subscription_period": {
                    "value": sub.subscription_period_value,
                    "unit": sub.subscription_period_unit
                },
                "effective_date": sub.effective_date,
                "billing_period": {
                    "value": sub.billing_period_value,
                    "unit": sub.billing_period_unit
                },
                "due_period": {
                    "value": sub.due_period_value or 1,
                    "unit": sub.due_period_unit or "months"
                },
            }
            due_subscriptions.append(subscription_dict)
    
    if len(due_subscriptions) == 0:
        return DueTomorrowResponse(message="No subscriptions due tomorrow")
    return DueTomorrowListResponse(subscriptions=due_subscriptions)

@app.delete("/recurring-expenses/{subscription_id}")
def delete_recurring_expense(subscription_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    subscription = db.query(models.RecurringExpense).filter(models.RecurringExpense.id == subscription_id, models.RecurringExpense.user_id == current_user.id).first()
    if subscription is None:
        raise HTTPException(status_code=404, detail="Subscription not found or not authorized")
    
    message = f"Subscription {subscription.name} amounting to {subscription.amount} deleted successfully"
    db.delete(subscription)
    db.commit()
    return message

@app.get("/expenses/intention-breakdown")
def get_intention_breakdown(month: int, year: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # Get all expenses for the specified month and year for the current user
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    expenses = db.query(models.Expense).filter(
        models.Expense.date >= start_date,
        models.Expense.date < end_date,
        models.Expense.user_id == current_user.id
    ).all()
    
    # Calculate totals for each intention
    intention_totals = {
        "Need": 0,
        "Want": 0,
        "Saving": 0
    }
    
    for expense in expenses:
        intention_totals[expense.intention] += expense.amount
    
    # Calculate percentages
    total = sum(intention_totals.values())
    intention_percentages = {
        intention: (amount / total * 100) if total > 0 else 0
        for intention, amount in intention_totals.items()
    }
    
    return {
        "totals": intention_totals,
        "percentages": intention_percentages,
        "total_amount": total
    }

@app.post("/budget/", response_model=Budget)
def create_budget(budget: BudgetCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # Convert category names to IDs
    category_budgets = {}
    for category_name, amount in budget.category_budgets.items():
        category_id = next((k for k, v in CATEGORIES.items() if v == category_name), None)
        if category_id is None:
            raise HTTPException(status_code=400, detail=f"Invalid category: {category_name}")
        category_budgets[category_id] = amount

    db_budget = models.Budget(
        monthly_income=budget.monthly_income,
        saving_goal=budget.saving_goal,
        category_budgets=category_budgets,
        created_at=datetime.now(),
        user_id=current_user.id # Assign user_id
    )
    db.add(db_budget)
    db.commit()
    db.refresh(db_budget)
    
    # Convert category IDs back to names for response
    response_budgets = {}
    for cat_id, amount in db_budget.category_budgets.items():
        response_budgets[CATEGORIES[int(cat_id)]] = amount

    
    return Budget(
        id=db_budget.id,
        monthly_income=db_budget.monthly_income,
        saving_goal=db_budget.saving_goal,
        category_budgets=response_budgets,
        created_at=db_budget.created_at
    )

@app.get("/budget/latest", response_model=Budget)
def get_latest_budget(month: int = None, year: int = None, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    query = db.query(models.Budget).filter(models.Budget.user_id == current_user.id)
    
    if month is not None and year is not None:
        # Filter budgets for the specified month and year
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        query = query.filter(
            models.Budget.created_at >= start_date,
            models.Budget.created_at < end_date
        )
    
    budget = query.order_by(models.Budget.created_at.desc()).first()
    if not budget:
        raise HTTPException(status_code=200, detail="No budget found")
    
    # Convert category IDs to names
    response_budgets = {}
    for cat_id, amount in budget.category_budgets.items():
        response_budgets[CATEGORIES[int(cat_id)]] = amount
    
    return Budget(
        id=budget.id,
        monthly_income=budget.monthly_income,
        saving_goal=budget.saving_goal,
        category_budgets=response_budgets,
        created_at=budget.created_at
    )

@app.put("/budget/latest", response_model=Budget)
def update_latest_budget(budget: BudgetCreate, month: int = None, year: int = None, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    query = db.query(models.Budget).filter(models.Budget.user_id == current_user.id)
    
    if month is not None and year is not None:
        # Filter budgets for the specified month and year
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        query = query.filter(
            models.Budget.created_at >= start_date,
            models.Budget.created_at < end_date
        )
    
    latest_budget = query.order_by(models.Budget.created_at.desc()).first()
    
    if not latest_budget:
        # If no budget exists for the specified month/year, create a new one
        category_budgets = {}
        for category_name, amount in budget.category_budgets.items():
            category_id = next((k for k, v in CATEGORIES.items() if v == category_name), None)
            if category_id is None:
                raise HTTPException(status_code=400, detail=f"Invalid category: {category_name}")
            category_budgets[int(category_id)] = amount

        new_budget = models.Budget(
            monthly_income=budget.monthly_income,
            saving_goal=budget.saving_goal,
            category_budgets=category_budgets,
            created_at=datetime.now(),
            user_id=current_user.id # Assign user_id
        )
        db.add(new_budget)
        db.commit()
        db.refresh(new_budget)
        latest_budget = new_budget
    else:
        # Update existing budget
        category_budgets = {}
        for category_name, amount in budget.category_budgets.items():
            category_id = next((k for k, v in CATEGORIES.items() if v == category_name), None)
            if category_id is None:
                raise HTTPException(status_code=400, detail=f"Invalid category: {category_name}")
            category_budgets[int(category_id)] = amount

        latest_budget.monthly_income = budget.monthly_income
        latest_budget.saving_goal = budget.saving_goal
        latest_budget.category_budgets = category_budgets
        latest_budget.created_at = datetime.now()

        db.commit()
        db.refresh(latest_budget)
    
    # Convert category IDs back to names for response
    response_budgets = {}
    for cat_id, amount in latest_budget.category_budgets.items():
        response_budgets[CATEGORIES[int(cat_id)]] = amount
    
    return Budget(
        id=latest_budget.id,
        monthly_income=latest_budget.monthly_income,
        saving_goal=latest_budget.saving_goal,
        category_budgets=response_budgets,
        created_at=latest_budget.created_at
    )

@app.get("/expenses/daily")
def get_daily_expenses(month: int, year: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # Calculate start and end dates for the month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    # Query expenses for the month and group by date for the current user
    daily_expenses = db.query(
        models.Expense.date,
        func.sum(models.Expense.amount).label('total')
    ).filter(
        models.Expense.date >= start_date,
        models.Expense.date < end_date,
        models.Expense.user_id == current_user.id
    ).group_by(
        models.Expense.date
    ).order_by(
        models.Expense.date
    ).all()
    
    # Format the response
    return {
        "daily_expenses": [
            {
                "date": expense.date.strftime("%Y-%m-%d"),
                "amount": float(expense.total)
            }
            for expense in daily_expenses
        ]
    }

# Income APIs
@app.post("/income/", response_model=Income, tags=["Income"])
def create_income(income: IncomeCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    if income.category_id not in INCOME_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category ID. Must be between 1 and {len(INCOME_CATEGORIES)}")
    
    db_income = models.Income(**income.dict(), user_id=current_user.id)
    db.add(db_income)
    db.commit()
    db.refresh(db_income)
    return Income(**db_income.__dict__)

@app.get("/income/", response_model=List[Income], tags=["Income"])
def read_incomes(month: int = None, year: int = None, skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # If month and year are provided, filter incomes for that month
    query = db.query(models.Income).filter(models.Income.user_id == current_user.id)
    if month is not None and year is not None:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        query = query.filter(
            models.Income.date >= start_date,
            models.Income.date < end_date
        )
    # Sort by date first, then by created_at for consistent ordering
    incomes = query.order_by(models.Income.date.desc(), models.Income.created_at.desc()).offset(skip).limit(limit).all()
    return [Income(**income.__dict__) for income in incomes]

@app.delete("/income/{income_id}", tags=["Income"])
def delete_income(income_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    income = db.query(models.Income).filter(models.Income.id == income_id, models.Income.user_id == current_user.id).first()
    if income is None:
        raise HTTPException(status_code=404, detail="Income not found or not authorized")
    
    message = f"Income on {income.date} for {INCOME_CATEGORIES[income.category_id]} amounting to {income.amount} deleted successfully"
    db.delete(income)
    db.commit()
    return {"message": message}

@app.get("/income/total", tags=["Income"])
def get_total_income(month: int = None, year: int = None, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # If month and year are provided, filter incomes for that month
    query = db.query(models.Income).filter(models.Income.user_id == current_user.id)
    if month is not None and year is not None:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        # Get overall total for the month
        total = query.filter(
            models.Income.date >= start_date,
            models.Income.date < end_date
        ).with_entities(func.sum(models.Income.amount)).scalar()
        
        # Get category-wise totals for the month
        category_totals = query.with_entities(
            models.Income.category_id,
            func.sum(models.Income.amount).label('total')
        ).filter(
            models.Income.date >= start_date,
            models.Income.date < end_date
        ).group_by(models.Income.category_id).all()
    else:
        # Get overall total
        total = query.with_entities(func.sum(models.Income.amount)).scalar()
        
        # Get category-wise totals
        category_totals = query.with_entities(
            models.Income.category_id,
            func.sum(models.Income.amount).label('total')
        ).group_by(models.Income.category_id).all()
    
    # Convert category IDs to names and format the response
    category_breakdown = {
        INCOME_CATEGORIES[cat_id]: amount for cat_id, amount in category_totals
    }
    
    return {
        "overall_total": total if total else 0,
        "category_breakdown": category_breakdown
    }

# Saving APIs
@app.post("/savings/", response_model=Saving, tags=["Savings"])
def create_saving(saving: SavingCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    if saving.category_id not in SAVING_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category ID. Must be between 1 and {len(SAVING_CATEGORIES)}")
    
    db_saving = models.Saving(**saving.dict(), user_id=current_user.id)
    db.add(db_saving)
    db.commit()
    db.refresh(db_saving)
    return Saving(**db_saving.__dict__)

@app.get("/savings/", response_model=List[Saving], tags=["Savings"])
def read_savings(month: int = None, year: int = None, skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # If month and year are provided, filter savings for that month
    query = db.query(models.Saving).filter(models.Saving.user_id == current_user.id)
    if month is not None and year is not None:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        query = query.filter(
            models.Saving.date >= start_date,
            models.Saving.date < end_date
        )
    # Sort by date first, then by created_at for consistent ordering
    savings = query.order_by(models.Saving.date.desc(), models.Saving.created_at.desc()).offset(skip).limit(limit).all()
    return [Saving(**saving.__dict__) for saving in savings]

@app.delete("/savings/{saving_id}", tags=["Savings"])
def delete_saving(saving_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    saving = db.query(models.Saving).filter(models.Saving.id == saving_id, models.Saving.user_id == current_user.id).first()
    if saving is None:
        raise HTTPException(status_code=404, detail="Saving not found or not authorized")
    
    message = f"Saving on {saving.date} for {SAVING_CATEGORIES[saving.category_id]} amounting to {saving.amount} deleted successfully"
    db.delete(saving)
    db.commit()
    return {"message": message}

@app.get("/savings/total", tags=["Savings"])
def get_total_savings(month: int = None, year: int = None, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # If month and year are provided, filter savings for that month
    query = db.query(models.Saving).filter(models.Saving.user_id == current_user.id)
    if month is not None and year is not None:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        # Get overall total for the month (exclude "Saving Goal Redeemed" category)
        total = query.filter(
            models.Saving.date >= start_date,
            models.Saving.date < end_date,
            models.Saving.category_id != 8  # Exclude "Saving Goal Redeemed" category
        ).with_entities(func.sum(models.Saving.amount)).scalar()
        
        # Get category-wise totals for the month (exclude "Saving Goal Redeemed" category)
        category_totals = query.with_entities(
            models.Saving.category_id,
            func.sum(models.Saving.amount).label('total')
        ).filter(
            models.Saving.date >= start_date,
            models.Saving.date < end_date,
            models.Saving.category_id != 8  # Exclude "Saving Goal Redeemed" category
        ).group_by(models.Saving.category_id).all()
    else:
        # Get overall total (exclude "Saving Goal Redeemed" category)
        total = query.filter(
            models.Saving.category_id != 8  # Exclude "Saving Goal Redeemed" category
        ).with_entities(func.sum(models.Saving.amount)).scalar()
        
        # Get category-wise totals (exclude "Saving Goal Redeemed" category)
        category_totals = query.with_entities(
            models.Saving.category_id,
            func.sum(models.Saving.amount).label('total')
        ).filter(
            models.Saving.category_id != 8  # Exclude "Saving Goal Redeemed" category
        ).group_by(models.Saving.category_id).all()
    
    # Convert category IDs to names and format the response
    category_breakdown = {
        SAVING_CATEGORIES[cat_id]: amount for cat_id, amount in category_totals
    }
    
    return {
        "overall_total": total if total else 0,
        "category_breakdown": category_breakdown
    }

# Account APIs
@app.post("/accounts/", response_model=Account, tags=["Accounts"])
def create_account(account: AccountCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # Check if user already has an account
    existing_account = db.query(models.Account).filter(models.Account.user_id == current_user.id).first()
    if existing_account:
        raise HTTPException(status_code=400, detail="User already has an account. Use PUT to update balance.")
    
    db_account = models.Account(**account.dict(), user_id=current_user.id)
    db.add(db_account)
    db.commit()
    db.refresh(db_account)
    return Account(**db_account.__dict__)

@app.get("/accounts/", response_model=List[Account], tags=["Accounts"])
def get_accounts(db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    accounts = db.query(models.Account).filter(models.Account.user_id == current_user.id).all()
    return [Account(**account.__dict__) for account in accounts]

@app.put("/accounts/{account_id}", response_model=Account, tags=["Accounts"])
def update_account_balance(account_id: int, account: AccountCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_account = db.query(models.Account).filter(models.Account.id == account_id, models.Account.user_id == current_user.id).first()
    if not db_account:
        raise HTTPException(status_code=404, detail="Account not found or not authorized")
    
    db_account.balance = account.balance
    db_account.modified_at = datetime.now()
    db.commit()
    db.refresh(db_account)
    return Account(**db_account.__dict__)

# Monthly Summary API
@app.get("/monthly-summary", tags=["Summary"])
def get_monthly_summary(month: int, year: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    # Calculate start and end dates for the month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    # Get income totals
    income_total = db.query(func.sum(models.Income.amount)).filter(
        models.Income.user_id == current_user.id,
        models.Income.date >= start_date,
        models.Income.date < end_date
    ).scalar() or 0
    
    # Get expense totals
    expense_total = db.query(func.sum(models.Expense.amount)).filter(
        models.Expense.user_id == current_user.id,
        models.Expense.date >= start_date,
        models.Expense.date < end_date
    ).scalar() or 0
    
    # Get saving totals
    saving_total = db.query(func.sum(models.Saving.amount)).filter(
        models.Saving.user_id == current_user.id,
        models.Saving.date >= start_date,
        models.Saving.date < end_date
    ).scalar() or 0
    
    # Calculate net balance for the month (income - expenses - savings)
    net_balance = income_total - expense_total - saving_total
    
    # Get category breakdown for expenses
    expense_categories = db.query(
        models.Expense.category_id,
        func.sum(models.Expense.amount).label('total')
    ).filter(
        models.Expense.user_id == current_user.id,
        models.Expense.date >= start_date,
        models.Expense.date < end_date
    ).group_by(models.Expense.category_id).all()
    
    # Get category breakdown for income
    income_categories = db.query(
        models.Income.category_id,
        func.sum(models.Income.amount).label('total')
    ).filter(
        models.Income.user_id == current_user.id,
        models.Income.date >= start_date,
        models.Income.date < end_date
    ).group_by(models.Income.category_id).all()
    
    # Get category breakdown for savings (exclude "Saving Goal Redeemed" category)
    saving_categories = db.query(
        models.Saving.category_id,
        func.sum(models.Saving.amount).label('total')
    ).filter(
        models.Saving.user_id == current_user.id,
        models.Saving.date >= start_date,
        models.Saving.date < end_date,
        models.Saving.category_id != 8  # Exclude "Saving Goal Redeemed" category
    ).group_by(models.Saving.category_id).all()
    
    return {
        "month": month,
        "year": year,
        "totals": {
            "income": float(income_total),
            "expenses": float(expense_total),
            "savings": float(saving_total),
            "net_balance": float(net_balance)
        },
        "breakdowns": {
            "expenses": {
                CATEGORIES[cat_id]: float(amount) for cat_id, amount in expense_categories
            },
            "income": {
                INCOME_CATEGORIES[cat_id]: float(amount) for cat_id, amount in income_categories
            },
            "savings": {
                SAVING_CATEGORIES[cat_id]: float(amount) for cat_id, amount in saving_categories
            }
        }
    }

@app.post("/send-monthly-report")
def send_monthly_report(background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    today = date.today()
    first_day_of_current_month = today.replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    year = last_day_of_previous_month.year
    month = last_day_of_previous_month.month
    print(f"Sending monthly report for {month}/{year} to user {current_user.email}")

    report_data = _send_monthly_report_logic(db, year, month, current_user.id)
    if report_data:
        background_tasks.add_task(send_email, current_user.email, report_data["subject"], report_data["body"], is_html=True)
        return {"message": "Monthly report has been queued."}    
    return {"message": "No report was generated."}

@app.get("/monthly-report-html", response_class=HTMLResponse)
def get_monthly_report_html(month: int, year: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    report_data = _send_monthly_report_logic(db, year, month, current_user.id)
    if report_data:
        return HTMLResponse(content=report_data["body"])
    raise HTTPException(status_code=404, detail="Monthly report not found or could not be generated.")

class SavingGoalBase(BaseModel):
    name: str
    target_date: date
    target_amount: float
    saved_amount: float = 0

    class Config:
        orm_mode = True

class SavingGoalCreate(SavingGoalBase):
    pass

class SavingGoal(SavingGoalBase):
    id: int
    status: str = "active"
    is_completed: bool = False
    redeemed_at: datetime | None = None
    created_at: datetime

class SavingGoalEdit(BaseModel):
    name: str | None = None
    target_date: date | None = None
    target_amount: float | None = None

    class Config:
        orm_mode = True

class AddAmountRequest(BaseModel):
    amount: float

class RedeemGoalRequest(BaseModel):
    pass


@app.post("/saving-goals/", response_model=SavingGoal)
def create_saving_goal(goal: SavingGoalCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_goal = models.SavingGoal(**goal.dict(), user_id=current_user.id)
    db.add(db_goal)
    db.commit()
    db.refresh(db_goal)
    
    # Create linked savings record if initial saved_amount > 0
    if db_goal.saved_amount > 0:
        savings_record = models.Saving(
            user_id=current_user.id,
            name=f"{db_goal.name}",
            amount=db_goal.saved_amount,
            date=datetime.now().date(),
            category_id=7,  # "Saving Goal" category
            created_at=datetime.now()
        )
        db.add(savings_record)
        db.commit()
    
    return db_goal

@app.get("/saving-goals/", response_model=List[SavingGoal])
def get_saving_goals(db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    goals = db.query(models.SavingGoal).filter(models.SavingGoal.user_id == current_user.id).all()
    
    # Calculate saved_amount from savings records for each goal
    for goal in goals:
        total_saved = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == current_user.id,
            models.Saving.category_id == 7,  # "Saving Goal" category
            models.Saving.name == f"{goal.name}"
        ).scalar() or 0
        
        goal.saved_amount = total_saved
        
        # Update completion status based on calculated amount
        if total_saved >= goal.target_amount:
            goal.is_completed = True
            if goal.status == "active":
                goal.status = "completed"
        else:
            goal.is_completed = False
            if goal.status == "completed":
                goal.status = "active"
    
    return goals

@app.put("/saving-goals/{goal_id}", response_model=SavingGoal)
def update_saving_goal(goal_id: int, goal: SavingGoalCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_goal = db.query(models.SavingGoal).filter(models.SavingGoal.id == goal_id, models.SavingGoal.user_id == current_user.id).first()
    if not db_goal:
        raise HTTPException(status_code=404, detail="Saving goal not found or not authorized")
    
    for key, value in goal.dict().items():
        setattr(db_goal, key, value)
    
    db.commit()
    db.refresh(db_goal)
    return db_goal

@app.post("/saving-goals/{goal_id}/add-amount", response_model=SavingGoal)
def add_amount_to_saving_goal(goal_id: int, request: AddAmountRequest, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_goal = db.query(models.SavingGoal).filter(models.SavingGoal.id == goal_id, models.SavingGoal.user_id == current_user.id).first()
    if not db_goal:
        raise HTTPException(status_code=404, detail="Saving goal not found or not authorized")

    if db_goal.status == "redeemed":
        raise HTTPException(status_code=400, detail="Cannot add amount to redeemed goal")

    # Create linked savings record
    savings_record = models.Saving(
        user_id=current_user.id,
        name=f"{db_goal.name}",
        amount=request.amount,
        date=datetime.now().date(),
        category_id=7,  # "Saving Goal" category
        created_at=datetime.now()
    )
    db.add(savings_record)
    
    # Calculate total saved amount from savings records
    total_saved = db.query(func.sum(models.Saving.amount)).filter(
        models.Saving.user_id == current_user.id,
        models.Saving.category_id == 7,  # "Saving Goal" category
        models.Saving.name == f"{db_goal.name}"
    ).scalar() or 0
    total_saved += request.amount  # Add the new amount
    
    db_goal.saved_amount = total_saved
    
    # Check if goal is completed
    if total_saved >= db_goal.target_amount:
        db_goal.is_completed = True
        db_goal.status = "completed"
    
    db.commit()
    db.refresh(db_goal)
    return db_goal

@app.put("/saving-goals/{goal_id}/edit", response_model=SavingGoal)
def edit_saving_goal(goal_id: int, goal_update: SavingGoalEdit, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_goal = db.query(models.SavingGoal).filter(models.SavingGoal.id == goal_id, models.SavingGoal.user_id == current_user.id).first()
    if not db_goal:
        raise HTTPException(status_code=404, detail="Saving goal not found or not authorized")
    
    if db_goal.status == "redeemed":
        raise HTTPException(status_code=400, detail="Cannot edit redeemed goal")
    
    # Store old name for updating linked records
    old_name = db_goal.name
    
    # Update goal fields
    update_data = goal_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_goal, field, value)
    
    # Check if goal completion status changed
    if db_goal.saved_amount >= db_goal.target_amount:
        db_goal.is_completed = True
        db_goal.status = "completed"
    else:
        db_goal.is_completed = False
        db_goal.status = "active"
    
    # Update linked savings records if name changed
    if goal_update.name and goal_update.name != old_name:
        linked_savings = db.query(models.Saving).filter(
            models.Saving.user_id == current_user.id,
            models.Saving.category_id == 7,  # "Saving Goal" category
            models.Saving.name == f"{old_name}"
        ).all()
        
        for saving in linked_savings:
            saving.name = f"{goal_update.name}"
    
    db.commit()
    db.refresh(db_goal)
    return db_goal

@app.post("/saving-goals/{goal_id}/redeem", response_model=SavingGoal)
def redeem_saving_goal(goal_id: int, request: RedeemGoalRequest, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_goal = db.query(models.SavingGoal).filter(models.SavingGoal.id == goal_id, models.SavingGoal.user_id == current_user.id).first()
    if not db_goal:
        raise HTTPException(status_code=404, detail="Saving goal not found or not authorized")
    
    if db_goal.status == "redeemed":
        raise HTTPException(status_code=400, detail="Goal already redeemed")
    
    # Update goal status (maintain is_completed flag for achieved goals)
    was_completed = db_goal.is_completed
    db_goal.status = "redeemed"
    db_goal.redeemed_at = datetime.now()
    # Keep is_completed flag intact for achieved goals
    
    # Update linked savings records category to "Saving Goal Redeemed"
    linked_savings = db.query(models.Saving).filter(
        models.Saving.user_id == current_user.id,
        models.Saving.category_id == 7,  # "Saving Goal" category
        models.Saving.name == f"{db_goal.name}"
    ).all()
    
    for saving in linked_savings:
        saving.category_id = 8  # "Saving Goal Redeemed" category
    
    # Get or create user account
    user_account = db.query(models.Account).filter(models.Account.user_id == current_user.id).first()
    if not user_account:
        user_account = models.Account(user_id=current_user.id, balance=0.0, created_at=datetime.now())
        db.add(user_account)
    
    # Add redeemed amount back to real balance
    user_account.balance += db_goal.saved_amount
    user_account.modified_at = datetime.now()
    
    db.commit()
    db.refresh(db_goal)
    return db_goal

@app.delete("/saving-goals/{goal_id}")
def delete_saving_goal(goal_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    db_goal = db.query(models.SavingGoal).filter(models.SavingGoal.id == goal_id, models.SavingGoal.user_id == current_user.id).first()
    if not db_goal:
        raise HTTPException(status_code=404, detail="Saving goal not found or not authorized")

    # Delete linked savings records
    linked_savings = db.query(models.Saving).filter(
        models.Saving.user_id == current_user.id,
        models.Saving.category_id.in_([7, 8]),  # "Saving Goal" and "Saving Goal Redeemed" categories
        models.Saving.name == f"{db_goal.name}"
    ).all()
    
    for saving in linked_savings:
        db.delete(saving)
    
    db.delete(db_goal)
    db.commit()
    return {"message": "Saving goal and linked records deleted successfully"}

@app.post("/auth/register", response_model=User)
def register_user(user: UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = auth_service.get_password_hash(user.password)
    db_user = models.User(email=user.email, hashed_password=hashed_password, name=user.name, created_at=datetime.now())
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.post("/auth/token", response_model=TokenWithRefresh)
def login_for_access_token(response: Response, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not auth_service.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=auth_service.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth_service.create_access_token(
        data={"sub": str(user.id)}, expires_delta=access_token_expires
    )
    
    # Create refresh token
    refresh_token = auth_service.create_refresh_token(db, user.id)
    
    # Set refresh token as HttpOnly cookie
    response.set_cookie(
        key="refresh_token",
        value=refresh_token,
        max_age=auth_service.REFRESH_TOKEN_EXPIRE_DAYS * 24 * 60 * 60,  # Convert days to seconds
        httponly=True,
        secure=True,  # Use only over HTTPS in production
        samesite="lax"
    )
    
    # Update last login time
    user.last_login = datetime.now()
    db.commit()
    
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/auth/refresh", response_model=Token, tags=["Authentication"])
def refresh_access_token(request: Request, response: Response, db: Session = Depends(get_db)):
    """
    Refresh access token using refresh token from HttpOnly cookie.
    Implements token rotation - returns new access token and sets new refresh token cookie.
    """
    refresh_token = request.cookies.get("refresh_token")
    if not refresh_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Refresh token not found"
        )
    
    # Validate the refresh token and get user
    user = auth_service.validate_refresh_token(db, refresh_token)
    
    # Revoke the old refresh token (token rotation)
    auth_service.revoke_refresh_token(db, refresh_token)
    
    # Create new access token
    access_token_expires = timedelta(minutes=auth_service.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth_service.create_access_token(
        data={"sub": str(user.id)}, expires_delta=access_token_expires
    )
    
    # Create new refresh token (token rotation)
    new_refresh_token = auth_service.create_refresh_token(db, user.id)
    
    # Set new refresh token as HttpOnly cookie
    response.set_cookie(
        key="refresh_token",
        value=new_refresh_token,
        max_age=auth_service.REFRESH_TOKEN_EXPIRE_DAYS * 24 * 60 * 60,
        httponly=True,
        secure=True,
        samesite="lax"
    )
    
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/auth/logout", tags=["Authentication"])
def logout(request: Request, response: Response, db: Session = Depends(get_db), current_user: models.User = Depends(auth_service.get_current_user)):
    """
    Logout user by revoking all refresh tokens and clearing cookies.
    """
    # Revoke all refresh tokens for the user
    auth_service.revoke_all_user_refresh_tokens(db, current_user.id)
    
    # Clear the refresh token cookie
    response.delete_cookie(key="refresh_token", httponly=True, secure=True, samesite="lax")
    
    return {"message": "Successfully logged out"}

@app.post("/auth/request-password-reset", status_code=status.HTTP_200_OK)
def request_password_reset(request: RequestPasswordResetRequest, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == request.email).first()

    if user:
        token = auth_service.generate_password_reset_token(db, user.id)
        reset_link = f"http://localhost:5173/reset-password?token={token}" # Frontend reset password URL
        
        background_tasks.add_task(send_password_reset_email,
                                  user.email,
                                  reset_link)
    
    return {"message": "If an account with that email exists, a password reset link has been sent to your inbox."}

@app.post("/auth/reset-password", status_code=status.HTTP_200_OK)
def reset_password(request: ResetPasswordRequest, db: Session = Depends(get_db)):
    user = auth_service.get_user_from_password_reset_token(db, request.token)
    
    hashed_password = auth_service.get_password_hash(request.new_password)
    user.hashed_password = hashed_password
    db.commit()
    db.refresh(user)

    return {"message": "Your password has been reset successfully."}

@app.get("/auth/me", response_model=User, tags=["Authentication"])
def get_current_user_info(current_user: models.User = Depends(auth_service.get_current_user)):
    """Get current authenticated user information"""
    return User(
        id=current_user.id,
        email=current_user.email,
        name=current_user.name,
        created_at=current_user.created_at,
        last_login=current_user.last_login
    )

